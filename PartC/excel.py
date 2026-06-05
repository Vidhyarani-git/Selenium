#write the same data 
# import openpyxl
# file=r"C:\Users\Saravanan\Desktop\selenium\PartC\Test_excel1.xlsx"
# workbook=openpyxl.load_workbook(file)
# #print(workbook.sheetnames) #display active sheetnames
#
# sheet=workbook['Sheet1']
# rows=3
# columns=3
# for row in range(1,rows+1):
#     for column in range(1,columns+1):
#         sheet.cell(row,column).value="Test"
# workbook.save(file)

#write different data
# import openpyxl
#
# file=r"C:\Users\Saravanan\Desktop\selenium\PartC\Test_excel1.xlsx"
#
# workbook=openpyxl.load_workbook(file)
# sheet=workbook['Sheet2']
# sheet.cell(1,1).value="S.no"
# sheet.cell(1,2).value="Name"
# sheet.cell(1,3).value="EmpId"
#
# sheet.cell(2,1).value="01"
# sheet.cell(2,2).value="Sam"
# sheet.cell(2,3).value="E001"
#
# sheet.cell(3,1).value="02"
# sheet.cell(3,2).value="John"
# sheet.cell(3,3).value="E002"
# workbook.save(file)

#Read the data
# import openpyxl
# file=r"C:\Users\Saravanan\Desktop\selenium\PartC\Test_excel1.xlsx"
# workbook=openpyxl.load_workbook(file)
# sheet=workbook['Sheet2']
# row_num=sheet.max_row
# col_num=sheet.max_column
# for row in range(1,row_num+1):
#     for column in range(1,col_num+1):
#         print(sheet.cell(row,column).value,end=' ')
#     print()

#Fill the color
##import openpyxl
##from openpyxl.styles import PatternFill
##file=r"C:\Users\Saravanan\Desktop\selenium\PartC\Test_excel1.xlsx"
##workbook=openpyxl.load_workbook(file)
##sheet=workbook['Sheet2']
##green_color=PatternFill(fill_type='solid',start_color='ff0000',end_color='ff0000')
##red_color=PatternFill(fill_type='solid',start_color='00ff00',end_color='00ff00')
##sheet.cell(2,2).fill=red_color
##sheet.cell(3,2).fill=green_color
##workbook.save(file)



#Using Excel utilities
import openpyxl
from excel_utilities import *

file=r"C:\Users\Saravanan\Desktop\selenium\PartC\Test_excel1.xlsx"

# workbook=openpyxl.load_workbook(file)
# sheet=workbook["Sheet3"]

# rowNum=sheet.max_row
# colNum=sheet.max_column
# for row in range(1,rowNum+1):
#     for col in range(1,colNum+1):
#         print(readData(file,"Sheet2",row,col),end=' ')
#     print()

#write the data
# writeData(file,'Sheet3',1,1,'EmpId')
# writeData(file,'Sheet3',1,2,'EmpName')
# writeData(file,'Sheet3',1,3,'PhNo')
# writeData(file,'Sheet3',1,4,'Sal')
# writeData(file,'Sheet3',1,5,'Location')
#
# writeData(file,'Sheet3',2,1,'001')
# writeData(file,'Sheet3',2,2,'ABC')
# writeData(file,'Sheet3',2,3,'9876543210')
# writeData(file,'Sheet3',2,4,'23000')
# writeData(file,'Sheet3',2,5,'Tumkur')

writeData(file,'Sheet3',3,1,'002')
writeData(file,'Sheet3',3,2,'DEF')
writeData(file,'Sheet3',3,3,'876543210')
writeData(file,'Sheet3',3,4,'25000')
writeData(file,'Sheet3',3,5,'Hosa road')




