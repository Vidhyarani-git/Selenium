import openpyxl

def rowNum(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.max_row

def colNum(file, sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.max_column

def readData(file,sheetname,row,col):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.cell(row,col).value

def writeData(file,sheetname,row,col,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    sheet.cell(row,col).value=data
    workbook.save(file)

#change font bold style function
from openpyxl.styles import Font
def fillFont(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    # cell = sheet.cell(row=row, column=col)
    bold_font=Font(bold=True,color="FF000000")
    for cell in sheet[1]:
        cell.font=bold_font
    workbook.save(file)

#fill the color
from openpyxl.styles import PatternFill
def fillPattern(file,sheetname,row,col,Expected_result):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    #fill yellow color on header row
    # sheet.cell(row,col).fill=PatternFill(fill_type="solid",start_color="FFFF00",end_color="FFFF00")
    #fill green color and red color on expected result column
    green_color=PatternFill(fill_type="solid",start_color="00FF00",end_color="00FF00")
    red_color=PatternFill(fill_type="solid",start_color="FF0000",end_color="FF0000")
    # Check result in column 6
    # for row in range(2, sheet.max_row + 1):
    #     result = str(sheet.cell(row=row, column=6).value).lower()
    #     if result == "pass":
    #         sheet.cell(row=row, column=6).fill = green_color
    #     elif result == "fail":
    #         sheet.cell(row=row, column=6).fill = red_color
    # fill color on Actual result in column 7
    for row in range(2, sheet.max_row + 1):
        result = str(sheet.cell(row=row, column=7).value).lower()
        if result == "pass":
            sheet.cell(row=row, column=7).fill = green_color
        elif result == "fail":
            sheet.cell(row=row, column=7).fill = red_color
    workbook.save(file)


